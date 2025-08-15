import logging
import threading
from datetime import datetime, timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.http import Http404
from django.utils import timezone
from django.db import transaction
from django.db.models import Q, Count, Avg
from django.contrib import messages

from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse, JsonResponse
import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from .models import (
    ScrapingJob, Company, Director, CompanyContact, 
    ScrapingSource, ScrapingAttempt, ExportRequest
)
from .serializers import (
    ScrapingJobSerializer, ScrapingJobCreateSerializer, CompanySerializer,
    CompanyListSerializer, DirectorSerializer, CompanyContactSerializer,
    ScrapingSourceSerializer, ScrapingAttemptSerializer, ExportRequestSerializer,
    ExportRequestCreateSerializer, CompanySearchSerializer, ContactVerificationSerializer,
    BulkExportSerializer
)
from .services import ScrapingService, initialize_scraping_sources

logger = logging.getLogger('scraper')


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class ScrapingJobViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing scraping jobs
    """
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'created_at']
    search_fields = ['company_name']
    ordering_fields = ['created_at', 'completed_at', 'progress']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ScrapingJobCreateSerializer
        return ScrapingJobSerializer
    
    def get_queryset(self):
        return ScrapingJob.objects.filter(user=self.request.user).prefetch_related(
            'companies', 'attempts', 'companies__directors', 'companies__contacts'
        )
    
    def perform_create(self, serializer):
        job = serializer.save(user=self.request.user)
        
        # Start scraping in background thread
        def start_scraping():
            service = ScrapingService()
            service.start_scraping_job(job.id)
        
        thread = threading.Thread(target=start_scraping)
        thread.daemon = True
        thread.start()
        
        logger.info(f"Created scraping job {job.id} for user {self.request.user.username}")
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """
        Cancel a scraping job
        """
        job = self.get_object()
        
        if job.status in ['pending', 'in_progress']:
            job.status = 'cancelled'
            job.completed_at = timezone.now()
            job.save()
            
            return Response({'message': 'Job cancelled successfully'})
        else:
            return Response(
                {'error': 'Cannot cancel job with status: ' + job.status},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def restart(self, request, pk=None):
        """
        Restart a failed or cancelled job
        """
        job = self.get_object()
        
        if job.status in ['failed', 'cancelled']:
            job.status = 'pending'
            job.started_at = None
            job.completed_at = None
            job.progress = 0
            job.error_message = ''
            job.save()
            
            # Start scraping in background
            def start_scraping():
                service = ScrapingService()
                service.start_scraping_job(job.id)
            
            thread = threading.Thread(target=start_scraping)
            thread.daemon = True
            thread.start()
            
            return Response({'message': 'Job restarted successfully'})
        else:
            return Response(
                {'error': 'Cannot restart job with status: ' + job.status},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        Get user's scraping statistics
        """
        jobs = self.get_queryset()
        
        stats = {
            'total_jobs': jobs.count(),
            'completed_jobs': jobs.filter(status='completed').count(),
            'failed_jobs': jobs.filter(status='failed').count(),
            'in_progress_jobs': jobs.filter(status='in_progress').count(),
            'total_companies': Company.objects.filter(scraping_job__user=request.user).count(),
            'total_directors': Director.objects.filter(company__scraping_job__user=request.user).count(),
            'total_contacts': CompanyContact.objects.filter(company__scraping_job__user=request.user).count(),
        }
        
        # Recent activity (last 30 days)
        last_30_days = timezone.now() - timedelta(days=30)
        stats['recent_jobs'] = jobs.filter(created_at__gte=last_30_days).count()
        
        return Response(stats)


class CompanyViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing company data
    """
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['company_status', 'company_type', 'incorporation_date']
    search_fields = ['name', 'company_number', 'nature_of_business']
    ordering_fields = ['name', 'incorporation_date', 'created_at']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return CompanyListSerializer
        return CompanySerializer
    
    def get_queryset(self):
        return Company.objects.filter(
            scraping_job__user=self.request.user
        ).prefetch_related('directors', 'contacts', 'scraping_job')
    
    @action(detail=True, methods=['get'])
    def contacts(self, request, pk=None):
        """
        Get all contacts for a company
        """
        company = self.get_object()
        contacts = company.contacts.all()
        serializer = CompanyContactSerializer(contacts, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def directors(self, request, pk=None):
        """
        Get all directors for a company
        """
        company = self.get_object()
        directors = company.directors.all()
        serializer = DirectorSerializer(directors, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """
        Get only active companies
        """
        queryset = self.get_queryset().filter(
            company_status__iexact='active'
        )
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class DirectorViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing director data
    """
    serializer_class = DirectorSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['director_type', 'nationality', 'officer_role']
    search_fields = ['name', 'occupation']
    ordering_fields = ['name', 'appointed_on', 'created_at']
    ordering = ['-appointed_on']
    
    def get_queryset(self):
        return Director.objects.filter(
            company__scraping_job__user=self.request.user
        ).select_related('company')
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """
        Get only active directors (not resigned)
        """
        queryset = self.get_queryset().filter(resigned_on__isnull=True)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class CompanyContactViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing company contacts
    """
    serializer_class = CompanyContactSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['contact_type', 'source', 'verified']
    search_fields = ['contact_value']
    ordering_fields = ['confidence_score', 'found_at', 'verified']
    ordering = ['-confidence_score', '-found_at']
    
    def get_queryset(self):
        return CompanyContact.objects.filter(
            company__scraping_job__user=self.request.user
        ).select_related('company')
    
    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """
        Mark a contact as verified
        """
        contact = self.get_object()
        contact.verified = True
        contact.last_verified = timezone.now()
        contact.save()
        
        serializer = self.get_serializer(contact)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def verified(self, request):
        """
        Get only verified contacts
        """
        queryset = self.get_queryset().filter(verified=True)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_type(self, request):
        """
        Get contacts grouped by type
        """
        contact_type = request.query_params.get('type')
        if not contact_type:
            return Response(
                {'error': 'type parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(contact_type=contact_type)
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class ExportRequestViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing export requests
    """
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['format', 'status']
    ordering_fields = ['requested_at', 'completed_at']
    ordering = ['-requested_at']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ExportRequestCreateSerializer
        return ExportRequestSerializer
    
    def get_queryset(self):
        return ExportRequest.objects.filter(
            user=self.request.user
        ).prefetch_related('scraping_jobs')
    
    def perform_create(self, serializer):
        export_request = serializer.save(user=self.request.user)
        
        # TODO: Start export processing in background
        # This would typically involve creating files and storing them
        logger.info(f"Created export request {export_request.id} for user {self.request.user.username}")


class CompanySearchView(APIView):
    """
    Search for companies using the scraping service
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        serializer = CompanySearchSerializer(data=request.data)
        if serializer.is_valid():
            company_name = serializer.validated_data['company_name']
            max_results = serializer.validated_data.get('max_results', 10)
            
            # Use the scraping service to search
            service = ScrapingService()
            companies = service.companies_house_scraper.search_company(
                company_name, max_results
            )
            
            return Response({
                'query': company_name,
                'results': companies,
                'count': len(companies)
            })
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DashboardView(APIView):
    """
    Dashboard with overview statistics
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        
        # Get user's data
        jobs = ScrapingJob.objects.filter(user=user)
        companies = Company.objects.filter(scraping_job__user=user)
        directors = Director.objects.filter(company__scraping_job__user=user)
        contacts = CompanyContact.objects.filter(company__scraping_job__user=user)
        
        # Calculate statistics
        stats = {
            'overview': {
                'total_jobs': jobs.count(),
                'total_companies': companies.count(),
                'total_directors': directors.count(),
                'total_contacts': contacts.count(),
                'verified_contacts': contacts.filter(verified=True).count(),
            },
            'job_status': {
                'pending': jobs.filter(status='pending').count(),
                'in_progress': jobs.filter(status='in_progress').count(),
                'completed': jobs.filter(status='completed').count(),
                'failed': jobs.filter(status='failed').count(),
                'cancelled': jobs.filter(status='cancelled').count(),
            },
            'company_status': {
                'active': companies.filter(company_status__iexact='active').count(),
                'dissolved': companies.filter(company_status__iexact='dissolved').count(),
            },
            'contact_types': {},
            'recent_activity': []
        }
        
        # Contact types distribution
        contact_types = contacts.values('contact_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        for contact_type in contact_types:
            stats['contact_types'][contact_type['contact_type']] = contact_type['count']
        
        # Recent activity (last 10 jobs)
        recent_jobs = jobs.order_by('-created_at')[:10]
        for job in recent_jobs:
            stats['recent_activity'].append({
                'id': job.id,
                'company_name': job.company_name,
                'status': job.status,
                'created_at': job.created_at,
                'progress': job.progress
            })
        
        return Response(stats)





def home(request):
    """Home page view"""
    return render(request, 'home.html')


def search_companies(request):
    """Search and scrape company data"""
    import csv
    import io
    from django.http import HttpResponse, JsonResponse
    from django.contrib import messages
    
    if request.method == 'POST':
        company_name = request.POST.get('company_name', '').strip()
        
        if not company_name:
            messages.error(request, 'Please enter a company name')
            return render(request, 'home.html')
        
        # Check if this is an AJAX request for progress updates
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'searching',
                'progress': 0,
                'message': 'Starting search...'
            })
        
        try:
            # Create a scraping service instance
            service = ScrapingService()
            
            # Get consolidated company data with all information interlinked - max 50 results
            consolidated_data = service.get_consolidated_company_data(company_name, max_results=50)
            
            if not consolidated_data:
                return render(request, 'results.html', {
                    'company_name': company_name,
                    'consolidated_data': [],
                    'total_companies': 0,
                    'total_directors': 0,
                    'total_contacts': 0,
                    'verified_contacts_count': 0
                })
            
            # Calculate totals for summary
            total_companies = len(consolidated_data)
            total_directors = sum(len(company['directors']) for company in consolidated_data)
            total_contacts = sum(
                len(company['company_contacts']) + 
                sum(len(director['contacts']) for director in company['directors'])
                for company in consolidated_data
            )
            verified_contacts_count = 0  # All contacts are estimated for now
            
            # Store data in session for export
            export_data = {
                'company_name': company_name,
                'companies': [],
                'directors': [],
                'contacts': []
            }
            
            for company in consolidated_data:
                # Company data for export
                export_data['companies'].append({
                    'name': company['company_details']['name'],
                    'company_number': company['company_details']['company_number'],
                    'company_status': company['company_details']['company_status'],
                    'company_type': company['company_details']['company_type'],
                    'incorporation_date': str(company['company_details']['incorporation_date']) if company['company_details']['incorporation_date'] else '',
                    'address': company['company_details']['registered_office_address'].get('full_address', '') if isinstance(company['company_details']['registered_office_address'], dict) else '',
                    'sic_codes': ', '.join(company['company_details']['sic_codes']) if company['company_details']['sic_codes'] else '',
                    'source': company.get('source', 'unknown'),
                    'industry': company['company_details'].get('industry', ''),
                    'size': company['company_details'].get('size', ''),
                    'funding': company['company_details'].get('funding', ''),
                    'employees': company['company_details'].get('employees', ''),
                    'founded': company['company_details'].get('founded', ''),
                    'credit_rating': company['company_details'].get('credit_rating', ''),
                    'revenue': company['company_details'].get('revenue', ''),
                    'domain': company['company_details'].get('domain', '')
                })
                
                # Director data for export
                for director in company['directors']:
                    export_data['directors'].append({
                        'name': director['name'],
                        'company': company['company_details']['name'],
                        'officer_role': director['details']['officer_role'],
                        'nationality': director['details']['nationality'],
                        'occupation': director['details']['occupation'],
                        'appointed_on': str(director['details']['appointed_on']) if director['details']['appointed_on'] else '',
                        'status': 'Active' if director['details']['is_active'] else 'Resigned'
                    })
                
                # Contact data for export
                for contact in company['company_contacts']:
                    export_data['contacts'].append({
                        'company': company['company_details']['name'],
                        'related_to': f"{company['company_details']['name']} (Company)",
                        'relationship': 'company',
                        'type': contact['type'],
                        'contact': contact['value'],
                        'source': contact['source'],
                        'confidence': contact['confidence'],
                        'verified': 'No'
                    })
                
                for director in company['directors']:
                    for contact in director['contacts']:
                        export_data['contacts'].append({
                            'company': company['company_details']['name'],
                            'related_to': f"{director['name']} (Director)",
                            'relationship': 'director',
                            'type': contact['type'],
                            'contact': contact['value'],
                            'source': contact['source'],
                            'confidence': contact['confidence'],
                            'verified': 'No'
                        })
            
            request.session['last_search_data'] = export_data
            
            return render(request, 'results.html', {
                'company_name': company_name,
                'consolidated_data': consolidated_data,
                'total_companies': total_companies,
                'total_directors': total_directors,
                'total_contacts': total_contacts,
                'verified_contacts_count': verified_contacts_count
            })
            
        except Exception as e:
            messages.error(request, f'Error searching companies: {str(e)}')
            logger.error(f"Search error: {str(e)}")
            return render(request, 'home.html')
    
    return render(request, 'home.html')


def search_progress(request):
    """AJAX endpoint for search progress updates"""
    if request.method == 'GET':
        company_name = request.GET.get('company_name', '')
        
        # Simulate progress updates for different sources
        progress_data = {
            'companies_house': {
                'status': 'searching',
                'progress': 15,
                'message': 'Searching Companies House database...'
            },
            'linkedin': {
                'status': 'searching', 
                'progress': 30,
                'message': 'Searching LinkedIn company profiles...'
            },
            'google': {
                'status': 'searching',
                'progress': 45,
                'message': 'Searching Google for additional information...'
            },
            'crunchbase': {
                'status': 'searching',
                'progress': 60,
                'message': 'Searching Crunchbase for funding data...'
            },
            'dandb': {
                'status': 'searching',
                'progress': 75,
                'message': 'Searching Dun & Bradstreet for credit information...'
            },
            'processing': {
                'status': 'processing',
                'progress': 90,
                'message': 'Processing and consolidating results...'
            },
            'complete': {
                'status': 'complete',
                'progress': 100,
                'message': 'Search completed!'
            }
        }
        
        return JsonResponse(progress_data)
    
    return JsonResponse({'error': 'Invalid request method'})


def export_data(request):
    """Export search results to CSV or Excel"""
    import csv
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl import Workbook
    
    format_type = request.GET.get('format', 'csv')
    company_name = request.GET.get('company_name', 'companies')
    
    # Get data from session
    search_data = request.session.get('last_search_data', {})
    
    if not search_data:
        messages.error(request, 'No data available for export. Please search first.')
        return redirect('/')
    
    if format_type == 'csv':
        return export_csv(search_data, company_name)
    elif format_type == 'xlsx':
        return export_excel(search_data, company_name)
    else:
        messages.error(request, 'Invalid export format')
        return redirect('/')


def export_csv(search_data, company_name):
    """Export data to CSV format matching the session data structure"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{company_name}_consolidated_data.csv"'
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [
        'Company Name', 'Source', 'Company Email', 'Company Phone', 
        'Company LinkedIn', 'Company Website',
    ]
    writer.writerow(headers)
    
    # Process session data structure
    companies = search_data.get('companies', [])
    directors = search_data.get('directors', [])
    contacts = search_data.get('contacts', [])
    
    # Create lookup dictionaries for easier access
    company_lookup = {company['name']: company for company in companies}
    director_lookup = {}
    for director in directors:
        company_name = director['company']
        if company_name not in director_lookup:
            director_lookup[company_name] = []
        director_lookup[company_name].append(director)
    
    # Group contacts by company and type
    company_contacts = {}
    
    for contact in contacts:
        company_name = contact['company']
        contact_type = contact['type']
        contact_value = contact['contact']
        
        if contact['relationship'] == 'company':
            if company_name not in company_contacts:
                company_contacts[company_name] = {'email': [], 'phone': [], 'linkedin': [], 'website': []}
            if contact_type in company_contacts[company_name]:
                company_contacts[company_name][contact_type].append(contact_value)
        
    # Write data rows
    for company in companies:
        company_name = company['name']
        source = company['source']
        
        # Get company contacts
        company_contact_data = company_contacts.get(company_name, {})
        company_emails = company_contact_data.get('email', [])
        company_phones = company_contact_data.get('phone', [])
        company_linkedin = company_contact_data.get('linkedin', [])
        company_website = company_contact_data.get('website', [])
        
        # Get directors for this company
        
       
        
        # No directors - just company info
        writer.writerow([
            company_name.split(' (')[0].strip(),
            source,
            '; '.join(company_emails) if company_emails else '',
            '; '.join(company_phones) if company_phones else '',
            '; '.join(company_linkedin) if company_linkedin else '',
            '; '.join(company_website) if company_website else '',
            '', '', '', ''
        ])

    response.write(output.getvalue())
    return response


def export_excel(search_data, company_name):
    """Export data to Excel format matching the session data structure"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Data"

    headers = [
        'Company Name', 'Source', 'Company Email', 'Company Phone', 
        'Company LinkedIn', 'Company Website'
    ]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill

    # Process session data structure
    companies = search_data.get('companies', [])
    contacts = search_data.get('contacts', [])
    
    # Create lookup dictionaries for easier access
    company_lookup = {company['name']: company for company in companies}
    
    # Group contacts by company and type
    company_contacts = {}
    
    for contact in contacts:
        company_name = contact['company']
        contact_type = contact['type']
        contact_value = contact['contact']
        
        if contact['relationship'] == 'company':
            if company_name not in company_contacts:
                company_contacts[company_name] = {'email': [], 'phone': [], 'linkedin': [], 'website': []}
            if contact_type in company_contacts[company_name]:
                company_contacts[company_name][contact_type].append(contact_value)
        
    # Write data rows
    for company in companies:
        company_name = company['name']
        source = company['source']
        
        # Get company contacts
        company_contact_data = company_contacts.get(company_name, {})
        company_emails = company_contact_data.get('email', [])
        company_phones = company_contact_data.get('phone', [])
        company_linkedin = company_contact_data.get('linkedin', [])
        company_website = company_contact_data.get('website', [])
        
        # Get directors for this company
        
       
        # No directors - just company info
        ws.append([
            company_name.split(' (')[0].strip(),
            source,
            '; '.join(company_emails) if company_emails else '',
            '; '.join(company_phones) if company_phones else '',
            '; '.join(company_linkedin) if company_linkedin else '',
            '; '.join(company_website) if company_website else '',
            '', '', '', ''
        ])

    # Auto column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{company_name}_consolidated_data.xlsx"'
    wb.save(response)
    return response

# Note: Scraping sources initialization is handled in management commands
